#!/usr/bin/env python3
"""
Enriched list: add independent electricians found via web research.
Focused on small artisans/independents near Arzier (within 100km).
Deduplicates against existing client list.
"""

import openpyxl, re

ref_path = r"C:\Users\openc\.openclaw\media\inbound\Clients_de_Safe_OIBT_Control---d0fcb6aa-77e8-42af-8c09-ff3014d6f8a7.xlsx"
wb_ref = openpyxl.load_workbook(ref_path)

existing = {}
for sheet in wb_ref.sheetnames:
    ws = wb_ref[sheet]
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, 2).value or "").strip().lower()
        contact = str(ws.cell(r, 3).value or "").strip().lower()
        if nom:
            existing[nom] = {"nom": ws.cell(r, 2).value, "contact": contact}

def normalize(s):
    s = s.lower().strip()
    s = re.sub(r'\b(sa|sarl|gmbh|ag|sàrl)\b', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def is_dup(nom, email=""):
    nb = normalize(nom)
    for ex in existing:
        eb = normalize(ex)
        if len(nb) > 3 and len(eb) > 3:
            if nb == eb or nb in eb or eb in nb:
                return True
        if email:
            ex_contact = str(existing[ex]["contact"] or "").lower().strip()
            if email.lower().strip() == ex_contact:
                return True
    return False

# New independents found
independants = [
    # Region La Côte / Nyon / Gland (0-20km)
    ("RL Electricité Sàrl", "Nyon 1260", "+41796655429", "", "https://rl-electricite.ch"),
    ("Afcotech SA", "Nyon 1260", "", "", "https://www.afcotech.ch"),
    ("Rindis & Cie Sàrl", "Nyon 1260", "", "", "https://www.rindlisbacher-pierre.ch"),
    ("SEIC (Services Industriels)", "Gland 1196", "+41223648282", "info@seic.ch", "https://seic.ch"),
    ("Hugo Dubois - Electricien Nyon", "Nyon 1260", "+41795213704", "", "https://hugodubois.ch/electricien/vaud/nyon/"),
    ("Hugo Dubois - Electricien Morges", "Morges 1110", "+41795213704", "", "https://hugodubois.ch/electricien/vaud/morges/"),

    # Morges / Aubonne / Rolle (10-30km)
    ("Philsam SA", "Echandens 1023", "+41217014636", "", "https://mesartisans.ch/suisse-romande/electricien-morges/"),
    ("Durussel SA Electricité", "Lausanne 1018", "+41216437323", "info@durussel.ch", "https://mesartisans.ch/suisse-romande/electricien-lausanne-durussel-sa/"),
    ("Martignetti SA", "Founex 1297", "", "", "https://mesartisans.ch/suisse-romande/electricien-vaud-martignetti-sa/"),
    ("Dubey Communication-Electricité SA", "Vevey 1800", "+41219220090", "infos@dubey-electricite.ch", "http://www.dubey-electricite.ch"),
    
    # Lausanne / Crissier / Renens (20-40km)
    ("Avenir Electricité SA", "Lausanne", "", "", ""),
    ("Elec 2000", "Lausanne", "", "", ""),
    ("Théraulaz Electricité SA", "Cugy", "+41217311010", "", ""),
    ("Buchs Electricité SA", "Jouxtens-Mézery 1008", "", "", ""),
    
    # Yverdon / Nord Vaudois (30-50km)
    ("VOé SA", "Yverdon 1400", "", "info@voe.ch", "https://www.voe.ch"),
    ("Bärtschi Electricité Sàrl", "Yverdon 1400", "", "", ""),
    ("Crestoelec Sàrl", "Yverdon 1400", "", "info@crestoelec.ch", "https://www.crestoelec.ch"),
    ("Enerbat SA", "Yverdon", "", "", ""),

    # Genève (15-30km)
    ("Securelec", "Genève", "+41223081620", "securelec@securelec.ch", "https://securelec.ch"),
    ("Electro-plus SA", "Genève", "", "", ""),
    ("Urs Scherrer Electricité SA", "Genève", "", "", ""),
    ("Büchi Electricité SA", "Genève", "", "", ""),
    ("Chabloz Electricité SA", "Genève", "", "", ""),
    ("Maillardo Electricité Sàrl", "Genève", "", "", ""),
    ("Truffet SA", "Genève", "", "", ""),

    # Fribourg (<100km)
    ("Holderélectric SA", "Fribourg", "", "", ""),
    ("Electro Pont SA", "Bulle 1630", "", "", ""),

    # Valais ouest (50-100km)
    ("Electro-Techniques AZ (Valais)", "Martigny 1920", "+41217999191", "info@electrotechniques.ch", "https://electrotechniques.ch"),
    ("Stucky SA", "Martigny", "", "", ""),
]

# Existing companies (from previous search) that passed dedup
existants = [
    ("3TECH SA", "Les Acacias 1227", "+41223081500", "compta@3tech.ch", "http://www.3tech.ch"),
    ("AC électricité Sàrl", "Satigny 1242", "+41225109920", "info@ac-electricite.ch", "http://www.ac-electricite.ch"),
    ("ALVAZZI GROUPE SA", "Plan-les-Ouates 1228", "+41223019045", "geneve@alvazzigroupe.com", ""),
    ("AM Electricité SA", "Troinex 1256", "+41227840170", "contact@amelectricite.ch", "http://www.amelectricite.ch"),
    ("Bondat SA", "Vernier 1214", "+41223061000", "info@bondat.ch", ""),
    ("Boymond Electricité SARL", "La Croix-de-Rozon 1257", "+41227711797", "admin@boymond.ch", "http://www.boymond.ch"),
    ("Cometel SA", "Genève 1205", "+41227080303", "info@cometel.ch", "http://www.cometel.ch"),
    ("Constantin Electricité", "Morges 1110", "+41218241436", "patrick@constantinelectricite.ch", "https://constantinelectricite.ch"),
    ("Bersier Electricité SA", "Préverenges 1028", "+41218031111", "info@bersierelect.ch", "https://www.bersierelect.ch"),
    ("AA Elec Controles SA", "Rolle", "+41215521075", "aa.elec@bluewin.ch", "https://www.aaeleccontroles.ch"),
    ("Dave Elec", "Le Mont-sur-Lausanne 1052", "+41219036367", "contact@dave-elec.ch", "https://www.dave-elec.ch"),
    ("Jordi SA", "Le Mont-sur-Lausanne 1052", "+41213103925", "courriel@jordisa.ch", "http://www.jordisa.ch"),
    ("Electro-Techniques AZ SA", "Lausanne/Grandvaux", "+41217999191", "info@electrotechniques.ch", "https://electrotechniques.ch"),
    ("Eltex Electro-Téléphone SA", "Genève 1202", "+41227338111", "support@eltex.ch", "http://www.eltexgeneve.ch"),
    ("Energia Electricité SA", "Genève 1203", "+41227930100", "info@energia-sa.ch", "http://www.energia-sa.ch"),
    ("J. Kappeler SA", "Grand-Lancy 1212", "+41227071540", "info@kappeler-electricien.ch", "http://www.kappeler-electricien.ch"),
    ("Jäggi Philippe", "Collonge-Bellerive 1245", "+41223404040", "info@jaeggi-electro.ch", "http://www.jaeggi-electro.ch"),
    ("Kreutzer et Cie SA", "Genève 1205", "+41228001414", "info@kreutzer-electricite.ch", "http://www.kreutzer-electricite.ch"),
    ("LumiA Electricité Sàrl", "Meyrin 1217", "+41798463051", "info@lumi-a.ch", ""),
    ("MP Electricité SA", "Bernex 1233", "+41227571002", "info@mp-electricite.ch", "http://www.mp-electricite.ch"),
    ("Paul Puchat SA", "Plan-les-Ouates 1228", "+41227949666", "info@puchatsa.ch", ""),
    ("Idmelectric Sàrl", "Carouge 1227", "+41783224737", "info@idmelectric.ch", ""),
    ("Gattuso Electricité Sàrl", "Plan-les-Ouates 1228", "+41786742995", "info@gattusoelectricite.ch", ""),
    ("PermaLife Energy Sàrl", "Carouge 1227", "+41783185474", "j.keller@permalife-energy.ch", "http://www.permalife-energy.ch"),
    ("Team Swiss Elec Sàrl", "Petit-Lancy 1213", "+41792796573", "info@teamswisselec.ch", ""),
    ("Genève Électricité SARL", "Genève 1203", "+41762384702", "info@geneveelectricite.ch", "https://geneveelectricite.ch"),
    ("Ciel Electricité SA - Succ. La Côte", "Nyon 1260", "+41216514555", "lacote@cielelectricite.com", "http://www.cielelectricite.com"),
    ("DA Electro", "Gland 1196", "+41786086588", "daelectro@bluewin.ch", "https://www.daelectro.ch"),
    ("RAMA OIBT", "Vich 1267", "+41215616577", "info@ramaoibt.ch", "https://ramaoibt.ch"),
]

all_prospects = []
seen_final = set()

for nom, lieu, tel, email, site in existants + independants:
    if not is_dup(nom, email):
        key = email.lower().strip() if email else normalize(nom)
        if key and key not in seen_final:
            seen_final.add(key)
            all_prospects.append((nom, lieu, tel, email, site))

print(f"Nouveaux prospects (total): {len(all_prospects)}")
print(f"Dont existants + independants")

# Write Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Electriciens"

headers = ["Nom entreprise", "Lieu / Code postal", "Telephone", "Email", "Site web", "Statut", "Type"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, (nom, lieu, tel, email, site) in enumerate(all_prospects, 2):
    ws.cell(i, 1, nom)
    ws.cell(i, 2, lieu)
    ws.cell(i, 3, tel)
    ws.cell(i, 4, email)
    ws.cell(i, 5, site)
    ws.cell(i, 6, "Nouveau prospect")
    ws.cell(i, 7, "SARL/SA" if any(x in nom.lower() for x in ["sa", "sarl", "gmbh", "ag"]) else "Independant/Artisan")

for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(all_prospects) + 2):
        val = ws.cell(row, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[chr(64 + col)].width = max_len + 2

output = r"C:\Users\openc\.openclaw\workspace838360131817\Nouveaux_Prospects_Electriciens.xlsx"
wb.save(output)
print(f"Fichier: {output}")

# Stats
indep = sum(1 for _,_,_,_,_ in all_prospects if "Independant" in headers[-1:])
print(f"\nStats finales:")
print(f"  Total nouveaux prospects: {len(all_prospects)}")
print(f"  Avec telephone: {sum(1 for _,_,t,_,_ in all_prospects if t)}")
print(f"  Avec email: {sum(1 for _,_,_,e,_ in all_prospects if e)}")
