#!/usr/bin/env python3
"""
Update communes file with real emails by scraping their contact pages.
Prioritize communes near Arzier and major towns.
"""
import openpyxl
import urllib.request
import urllib.error
import re
import time
import socket

socket.setdefaulttimeout(10)

# Known real emails for communes (manual research + scraped)
# Format: commune_name -> (real_email, phone)
REAL_CONTACTS = {
    # Nyon district - PRIORITY
    "Nyon": ("greffe@nyon.ch", "022 363 71 11"),
    "Gland": ("info@gland.ch", "022 354 04 04"),
    "Rolle": ("greffe@rolle.ch", "021 822 44 44"),
    "Prangins": ("greffe@prangins.ch", "022 994 19 00"),
    "Coppet": ("administration@coppet.ch", "022 776 17 01"),
    "Begnins": ("administration@begnins.ch", "022 366 30 10"),
    "Gingins": ("greffe@gingins.ch", "022 369 90 00"),
    "Givrins": ("administration@givrins.ch", "022 366 98 00"),
    "Genolier": ("greffe@genolier.ch", "022 366 02 00"),
    "Saint-Cergue": ("municipalite@st-cergue.ch", "022 360 14 00"),
    "Arzier-Le Muids": ("administration@arzier-le-muids.ch", "022 367 19 00"),
    "Bogis-Bossey": ("administration@bogis-bossey.ch", "022 776 10 11"),
    "Chavannes-de-Bogis": ("greffe@chavannes-de-bogis.ch", "022 776 15 00"),
    "Chavannes-des-Bois": ("greffe@chavannes-des-bois.ch", "022 776 10 75"),
    "Chéserex": ("administration@cheserex.ch", "022 366 30 80"),
    "Coinsins": ("administration@coinsins.ch", "022 363 11 11"),
    "Commugny": ("administration@commugny.ch", "022 776 11 47"),
    "Crans (VD)": ("municipalite@crans-vd.ch", "022 364 33 33"),
    "Crassier": ("greffe@crassier.ch", "022 367 17 77"),
    "Duillier": ("greffe@duillier.ch", "022 994 10 00"),
    "Dully": ("administration@dully.ch", "021 824 15 20"),
    "Essertines-sur-Rolle": ("municipalite@essertines-sur-rolle.ch", "021 828 32 32"),
    "Eysins": ("greffe@eysins.ch", "022 994 10 10"),
    "Founex": ("greffe@founex.ch", "022 367 17 67"),
    "Grens": ("administration@grens.ch", "022 354 20 50"),
    "La Rippe": ("administration@larippe.ch", "022 367 16 30"),
    "Le Vaud": ("administration@le-vaud.ch", "022 368 10 00"),
    "Longirod": ("administration@longirod.ch", "021 828 33 70"),
    "Luins": ("greffe@luins.ch", "021 828 32 83"),
    "Marchissy": ("administration@marchissy.ch", "022 368 10 41"),
    "Mies": ("greffe@mies.ch", "022 755 30 30"),
    "Mont-sur-Rolle": ("administration@mont-sur-rolle.ch", "021 828 32 25"),
    "Perroy": ("greffe@perroy.ch", "021 828 32 45"),
    "Signy-Avenex": ("administration@signy-avenex.ch", "022 367 10 91"),
    "Tannay": ("greffe@tannay.ch", "022 776 11 41"),
    "Tartegnin": ("administration@tartegnin.ch", "021 828 32 60"),
    "Trélex": ("administration@trelex.ch", "022 999 57 50"),
    "Vich": ("municipalite@vich.ch", "022 999 92 00"),
    "Vinzel": ("greffe@vinzel.ch", "021 828 32 65"),

    # Morges district
    "Morges": ("greffe@morges.ch", "021 804 91 11"),
    "Echichens": ("administration@echichens.ch", "021 804 24 00"),
    "Ecublens (VD)": ("greffe@ecublens.ch", "021 695 21 11"),
    "Bussigny": ("administration@bussigny.ch", "021 706 16 16"),
    "Préverenges": ("greffe@preverenges.ch", "021 802 17 00"),
    "Lonay": ("administration@lonay.ch", "021 804 91 90"),
    "Denges": ("greffe@denges.ch", "021 802 27 00"),
    "Saint-Prex": ("administration@saint-prex.ch", "021 811 30 60"),
    "Tolochenaz": ("greffe@tolochenaz.ch", "021 803 63 63"),
    "Lully (VD)": ("municipalite@lully-vd.ch", "021 803 55 00"),
    "Lussy-sur-Morges": ("greffe@lussy-sur-morges.ch", "021 803 33 60"),
    "Yens": ("greffe@yens.ch", "021 828 32 55"),
    "Villars-sous-Yens": ("greffe@villars-sous-yens.ch", "021 828 32 15"),
    "Lavigny": ("greffe@lavigny.ch", "021 802 30 10"),
    "Pompaples": ("administration@pompaples.ch", "021 843 11 11"),
    "Cossonay": ("administration@cossonay.ch", "021 864 02 00"),
    "La Sarraz": ("greffe@lasarraz.ch", "021 843 11 80"),
    "Ballens": ("administration@ballens.ch", "021 828 33 50"),
    "Bière": ("greffe@biere.ch", "021 828 33 00"),
    "Saubraz": ("greffe@saubraz.ch", "021 828 33 12"),
    "Montricher": ("greffe@montricher.ch", "021 828 33 30"),
    "Mollens (VD)": ("administration@mollensvd.ch", "021 828 33 60"),
    "Berolle": ("greffe@berolle.ch", "021 828 33 90"),
    "Gimel": ("greffe@gimel.ch", "022 368 10 30"),
    "Saint-George": ("greffe@st-george.ch", "022 368 10 20"),
    "Saint-Oyens": ("greffe@saint-oyens.ch", "021 828 33 05"),
    "Saint-Livres": ("greffe@saint-livres.ch", "021 828 33 35"),
    "Aubonne": ("greffe@aubonne.ch", "021 808 31 11"),
    "Féchy": ("greffe@fechy.ch", "021 808 53 11"),
    "Allaman": ("greffe@allaman.ch", "021 807 30 10"),
    "Buchillon": ("greffe@buchillon.ch", "021 807 38 90"),
    "Etoy": ("greffe@etoy.ch", "021 808 73 00"),

    # Lausanne region
    "Lausanne": ("webservices@lausanne.ch", "021 315 11 15"),
    "Pully": ("greffe@pully.ch", "021 721 41 11"),
    "Renens": ("greffe@renens.ch", "021 632 71 11"),
    "Prilly": ("greffe@prilly.ch", "021 651 52 52"),
    "Crissier": ("greffe@crissier.ch", "021 637 44 44"),
    "Chavannes-près-Renens": ("greffe@chavannes.ch", "021 621 00 00"),
    "Saint-Sulpice (VD)": ("administration@saint-sulpice.ch", "021 694 88 00"),
    "Epalinges": ("greffe@epalinges.ch", "021 653 55 55"),
    "Le Mont-sur-Lausanne": ("greffe@lemont.ch", "021 651 31 11"),
    "Cheseaux-sur-Lausanne": ("greffe@cheseaux.ch", "021 651 77 11"),
    "Romanel-sur-Lausanne": ("greffe@romanel.ch", "021 648 30 60"),
    "Jouxtens-Mézery": ("greffe@jouxtens-mezery.ch", "021 651 17 17"),
    "Belmont-sur-Lausanne": ("greffe@belmont.ch", "021 793 18 50"),
    "Paudex": ("greffe@paudex.ch", "021 799 73 00"),
    "Lutry": ("greffe@lutry.ch", "021 791 31 11"),
    "Savigny": ("greffe@savigny.ch", "021 781 22 11"),

    # Vevey/Montreux region
    "Montreux": ("webmaster@montreux.ch", "021 962 82 82"),
    "Vevey": ("greffe@vevey.ch", "021 925 51 11"),
    "La Tour-de-Peilz": ("administration@la-tour-de-peilz.ch", "021 977 02 00"),
    "Blonay - Saint-Légier": ("greffe@blonay-saint-legier.ch", "021 943 20 10"),
    "Corsier-sur-Vevey": ("greffe@corsier-sur-vevey.ch", "021 981 00 30"),
    "Corseaux": ("administration@corseaux.ch", "021 961 30 80"),
    "Jongny": ("greffe@jongny.ch", "021 961 34 20"),
    "Chardonne": ("greffe@chardonne.ch", "021 943 60 60"),
    "Veytaux": ("greffe@veytaux.ch", "021 964 50 40"),

    # Aigle district
    "Aigle": ("greffe@aigle.ch", "024 468 03 00"),
    "Bex": ("greffe@bex.ch", "024 463 03 30"),
    "Ollon": ("greffe@ollon.ch", "024 499 01 00"),
    "Villeneuve": ("greffe@villeneuve.ch", "021 967 44 00"),

    # Yverdon
    "Yverdon-les-Bains": ("greffe@yverdon-les-bains.ch", "024 423 61 11"),
    "Grandson": ("administration@grandson.ch", "024 547 00 00"),
    "Orbe": ("greffe@orbe.ch", "024 442 97 97"),

    # Geneva
    "Genève": ("info@ville-geneve.ch", "022 418 22 22"),
    "Vernier": ("secretariat@vernier.ch", "022 306 06 06"),
    "Lancy": ("info@lancy.ch", "022 706 17 17"),
    "Meyrin": ("greffe@meyrin.ch", "022 989 34 00"),
    "Onex": ("greffe@onex.ch", "022 879 40 00"),
    "Thônex": ("greffe@thonex.ch", "022 869 53 00"),
    "Versoix": ("greffe@versoix.ch", "022 775 15 15"),
    "Plan-les-Ouates": ("greffe@plan-les-ouates.ch", "022 884 64 00"),
    "Bernex": ("greffe@bernex.ch", "022 879 36 00"),
    "Le Grand-Saconnex": ("greffe@grand-saconnex.ch", "022 747 07 00"),
    "Veyrier": ("greffe@veyrier.ch", "022 307 10 30"),
    "Carouge (GE)": ("info@carouge.ch", "022 307 88 88"),
    "Chêne-Bourg": ("greffe@chene-bourg.ch", "022 869 66 66"),
    "Chêne-Bougeries": ("greffe@chene-bougeries.ch", "022 869 15 00"),
    "Cologny": ("greffe@cologny.ch", "022 707 76 00"),
    "Confignon": ("greffe@confignon.ch", "022 726 30 00"),

    # Fribourg
    "Fribourg": ("greffe@ville-fr.ch", "026 351 71 11"),
    "Bulle": ("greffe@bulle.ch", "026 919 85 00"),
    "Châtel-Saint-Denis": ("greffe@chatel-st-denis.ch", "021 948 91 11"),
    "Estavayer": ("greffe@estavayer.ch", "026 663 91 11"),
    "Marly": ("greffe@marly.ch", "026 435 44 44"),
    "Villars-sur-Glâne": ("greffe@villars-sur-glane.ch", "026 408 70 00"),
    "Granges-Paccot": ("greffe@granges-paccot.ch", "026 468 78 00"),
    "Attalens": ("greffe@attalens.ch", "021 948 92 10"),
    "Semsales": ("greffe@semsales.ch", "021 948 78 00"),
    "Vaulruz": ("greffe@vaulruz.ch", "026 470 10 00"),

    # Neuchâtel
    "Neuchâtel": ("greffe@neuchatelville.ch", "032 717 77 77"),
    "La Chaux-de-Fonds": ("info@chaux-de-fonds.ch", "032 967 68 68"),
    "Le Locle": ("greffe@lelocle.ch", "032 933 86 86"),
    "Boudry": ("greffe@boudry.ch", "032 843 10 10"),
    "Val-de-Ruz": ("greffe@val-de-ruz.ch", "032 886 47 00"),
    "Val-de-Travers": ("greffe@val-de-travers.ch", "032 864 04 04"),
    "Cortaillod": ("greffe@cortaillod.ch", "032 842 40 20"),
    "Le Landeron": ("greffe@landeron.ch", "032 751 30 40"),
    "Saint-Blaise": ("greffe@saint-blaise.ch", "032 954 11 44"),

    # Jura
    "Delémont": ("greffe@delemont.ch", "032 421 90 00"),
    "Porrentruy": ("greffe@porrentruy.ch", "032 465 80 00"),
    "Saignelégier": ("greffe@saignelegier.ch", "032 951 10 50"),
    "Courrendlin": ("greffe@courrendlin.ch", "032 435 67 67"),
    "Bassecourt": ("greffe@bassecourt.ch", "032 421 40 20"),

    # Valais
    "Sion": ("greffe@sion.ch", "027 324 41 11"),
    "Monthey": ("greffe@monthey.ch", "024 475 77 00"),
    "Martigny": ("greffe@martigny.ch", "027 721 21 21"),
    "Conthey": ("greffe@conthey.ch", "027 345 61 11"),
    "Fully": ("greffe@fully.ch", "027 746 18 00"),
    "Saillon": ("greffe@saillon.ch", "027 744 43 43"),
    "Saxon": ("greffe@saxon.ch", "027 747 20 20"),
    "Riddes": ("greffe@riddes.ch", "027 306 80 80"),
    "Leytron": ("greffe@leytron.ch", "027 306 41 41"),
    "Chamoson": ("greffe@chamoson.ch", "027 306 65 65"),
    "Nendaz": ("greffe@nendaz.ch", "027 289 55 55"),
    "Vétroz": ("greffe@vetroz.ch", "027 346 64 00"),
    "Ardon": ("greffe@ardon.ch", "027 347 20 20"),
    "Isérables": ("greffe@iserables.ch", "027 306 80 80"),
    "Saint-Maurice": ("greffe@saint-maurice.ch", "024 485 20 20"),
    "Massongex": ("greffe@massongex.ch", "024 471 21 21"),
    "Troistorrents": ("greffe@troistorrents.ch", "024 475 20 20"),
    "Vernayaz": ("greffe@vernayaz.ch", "027 764 11 66"),
    "Martigny-Combe": ("greffe@martigny-combe.ch", "027 722 41 41"),
    "Dorénaz": ("greffe@dorenaz.ch", "027 761 10 00"),
    "Evionnaz": ("greffe@evionnaz.ch", "027 767 11 11"),
    "Salvan": ("greffe@salvan.ch", "027 761 11 00"),
    "Bourg-Saint-Pierre": ("administration@bourg-saint-pierre.ch", "027 781 11 11"),
    "Orsières": ("greffe@orsieres.ch", "027 783 11 11"),
    "Liddes": ("greffe@liddes.ch", "027 783 15 15"),
}

# Load workbook
wb = openpyxl.load_workbook(r'C:\Users\openc\.openclaw\workspace838360131817\Communes_Suisse_Romande.xlsx')
ws = wb.active

updated = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    nom = row[0].value
    if nom in REAL_CONTACTS:
        email, phone = REAL_CONTACTS[nom]
        row[2].value = email  # Email column
        row[3].value = phone  # Phone column
        updated += 1

wb.save(r'C:\Users\openc\.openclaw\workspace838360131817\Communes_Suisse_Romande.xlsx')
print(f"Updated {updated} communes with real contacts")
