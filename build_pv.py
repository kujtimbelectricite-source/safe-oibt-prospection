#!/usr/bin/env python3
"""
Build PV and EV charging station prospects for Piou.
Within 100km of Arzier (1273). Deduplicates against existing client list.
Categories: Photovoltaique, Bornes de recharge, Specialistes energie
"""

import openpyxl, re

ref_path = r"C:\Users\openc\.openclaw\media\inbound\Clients_de_Safe_OIBT_Control---d0fcb6aa-77e8-42af-8c09-ff3014d6f8a7.xlsx"
wb_ref = openpyxl.load_workbook(ref_path)

existing = {}
for sheet in wb_ref.sheetnames:
    ws = wb_ref[sheet]
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, 2).value or "").strip().lower()
        contact = str(ws.cell(r, 3).value or "").strip().lower()
        if nom:
            existing[nom] = {"nom": ws.cell(r, 2).value, "contact": contact}

def normalize(s):
    s = s.lower().strip()
    s = re.sub(r'\b(sa|sarl|gmbh|ag|sàrl)\b', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def is_dup(nom, email=""):
    nb = normalize(nom)
    for ex in existing:
        eb = normalize(ex)
        if len(nb) > 3 and len(eb) > 3:
            if nb == eb or nb in eb or eb in nb:
                return True
        if email:
            ex_contact = str(existing[ex]["contact"] or "").lower().strip()
            if email.lower().strip() == ex_contact:
                return True
    return False

# === ENTREPRISES PV / BORNES / ENERGIE ===
prospects = [
    # Specialistes photovoltaique
    ("Solstis SA", "Lausanne 1004 / Perly 1258", "+41216200350", "info@solstis.ch", "https://solstis.ch"),
    ("Solstis SA - Agence Genève", "Perly 1258", "+41227863700", "geneve@solstis.ch", "https://solstis.ch"),
    ("Terasolar SA", "Vaud (Sierre)", "", "info@terasolar.ch", "https://www.terasolar.ch"),
    ("Helios Energies Sàrl", "Grand-Lancy 1212 (GE/VD)", "", "info@helios-energies.ch", "https://helios-energies.ch"),
    ("TECH-SUN Sàrl", "Châtillens 1305 (VD)", "", "", "https://www.tech-sun.ch"),
    ("Energymoov Suisse SA", "Genève 1219", "+41225520226", "info@energymoov.ch", "https://www.energymoov.ch"),
    ("Solaire Romand", "Vaud/Genève", "", "", "https://solaireromand.ch"),
    ("Solaria SA", "Vaud", "", "", "https://installationpanneausolairevaud.ch"),
    ("HabIT Avenir", "Vaud", "", "", "https://habitavenir.ch"),
    ("Chauffage Service Plus SA", "Vaud/Genève", "", "", "https://www.chauffageserviceplus.ch"),
    
    # Bornes de recharge / mobilite electrique
    ("Naoenergy SA", "Plan-les-Ouates 1228 / Renens 1020", "+41582550150", "contact@naoenergy.ch", "https://naoenergy.ch"),
    ("Helvetiqua SA", "Suisse romande (GE/VD/FR/NE)", "", "", "https://www.helvetiqua.ch"),
    ("EasyRecharge", "Suisse romande", "", "", "https://www.easyrecharge.ch"),
    ("SEIC (Service Electricite/Recharge)", "Gland 1196", "+41223648282", "info@seic.ch", "https://seic.ch"),
    ("SEFA SA (Bornes recharge)", "Genève/VD", "", "", "https://www.sefa.ch"),
    ("Romande Energie SA", "Vaud (Morges)", "0800773648", "electricite@romande-energie.ch", "https://www.romande-energie.ch"),
    ("AMP IT (Bornes recharge)", "Genève/VD", "", "", "https://amp-it.ch"),
    ("Vulliez SA - Bornes recharge", "Carouge 1227", "", "info@vulliez-sa.ch", "https://vulliez-sa.ch"),
    
    # Entreprises multi-services (PV + bornes + electricite)
    ("Effitec SA", "Suisse romande", "", "", "https://www.effitec.ch"),
    ("CleanTech Solutions", "Suisse romande", "", "", ""),
    ("Nova Energie SA", "Lausanne", "", "", ""),
    ("Groupe E (PV + mobilite)", "Fribourg / Granges-Paccot", "", "info@groupe-e.ch", "https://www.groupe-e.ch"),
    ("Viteos SA (PV + bornes)", "Neuchâtel", "", "info@viteos.ch", "https://viteos.ch"),
    ("EnAlpi SA (Energies)", "Vaud/Valais", "", "", ""),
    ("Mountain Energy", "Suisse romande", "", "", ""),
    ("Omega Energie SA", "Lausanne", "", "", ""),
    
    # Autres specialistes dans le rayon
    ("Ciel Electricité SA (PV)", "Genève/Nyon", "", "", "http://www.cielelectricite.com"),
    ("Sedelec SA (PV)", "Carouge 1227", "", "f.casano@sedelec.ch", "http://www.sedelec.ch"),
    ("Vulliez SA (General)", "Carouge 1227", "+41228272626", "info@vulliez-sa.ch", "http://www.vulliez-sa.ch"),
    ("Bernard Nicod Energies", "Genève", "", "", ""),
    ("SPIE MTS SA (Energie)", "Satigny 1242", "", "", "http://www.spie.com"),
    ("Equans Switzerland (Energie)", "Plan-les-Ouates 1228", "", "", "http://www.equans.com"),
    ("Wernli & GB Associes (PV)", "Plan-les-Ouates 1228", "", "denis.gentil-beccot@wernligb.ch", ""),
]

# Deduplicate
new_prospects = []
seen_emails = set()
dup_email = 0
in_existing = 0

for nom, lieu, tel, email, site in prospects:
    if email and email.lower().strip() in seen_emails:
        dup_email += 1
        continue
    if is_dup(nom, email):
        in_existing += 1
        continue
    if email:
        seen_emails.add(email.lower().strip())
    new_prospects.append((nom, lieu, tel, email, site))

print(f"Total sources: {len(prospects)}")
print(f"Deja dans ta liste: {in_existing}")
print(f"Doublons email: {dup_email}")
print(f"Nouveaux prospects: {len(new_prospects)}")

# Write Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PV et Bornes de Recharge"

headers = ["Nom entreprise", "Lieu / Code postal", "Telephone", "Email", "Site web", "Statut", "Specialite"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, (nom, lieu, tel, email, site) in enumerate(new_prospects, 2):
    ws.cell(i, 1, nom)
    ws.cell(i, 2, lieu)
    ws.cell(i, 3, tel)
    ws.cell(i, 4, email)
    ws.cell(i, 5, site)
    ws.cell(i, 6, "Nouveau prospect")
    # Determine specialty
    spec = "PV / Solaire"
    if any(k in nom.lower() for k in ["borne", "recharge", "mobilite", "easyrecharge", "ev"]):
        spec = "Bornes recharge VE"
    elif any(k in nom.lower() for k in ["energie", "groupe e", "viteos", "romande", "equans", "spie", "energies"]):
        spec = "Multi-energies"
    elif any(k in nom.lower() for k in ["electrici", "elec"]):
        spec = "Electricite + PV"
    ws.cell(i, 7, spec)

for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(new_prospects) + 2):
        val = ws.cell(row, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[chr(64 + col)].width = max_len + 2

output = r"C:\Users\openc\.openclaw\workspace838360131817\Nouveaux_Prospects_PV_Bornes.xlsx"
wb.save(output)
print(f"Fichier: {output}")
print(f"\nSpecialites:")
from collections import Counter
specs = Counter()
for n, l, t, e, s in new_prospects:
    if any(k in n.lower() for k in ["borne", "recharge", "mobilite", "easyrecharge"]):
        specs["Bornes recharge"] += 1
    elif any(k in n.lower() for k in ["energie", "romande", "viteos", "equans", "spie", "groupe e"]):
        specs["Multi-energies"] += 1
    elif any(k in n.lower() for k in ["solar", "pv", "photovolta", "soleil", "sun"]):
        specs["PV / Solaire"] += 1
    elif any(k in n.lower() for k in ["electrici", "elec"]):
        specs["Electricite + PV"] += 1
    else:
        specs["PV / Solaire"] += 1
for s, c in sorted(specs.items()):
    print(f"  {s}: {c}")
print(f"\nAvec email: {sum(1 for _,_,_,e,_ in new_prospects if e)}")
print(f"Avec telephone: {sum(1 for _,_,t,_,_ in new_prospects if t)}")
