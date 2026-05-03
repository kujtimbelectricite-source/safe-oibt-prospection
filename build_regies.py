#!/usr/bin/env python3
"""
Build regies & agences immobilières prospects for Piou.
Within 100km of Arzier (1273). Deduplicates against existing client list.
Categories: Regies immobilières, Agences immobilières, Gérance
"""

import openpyxl, re

ref_path = r"C:\Users\openc\.openclaw\media\inbound\Clients_de_Safe_OIBT_Control---d0fcb6aa-77e8-42af-8c09-ff3014d6f8a7.xlsx"
wb_ref = openpyxl.load_workbook(ref_path)

existing = {}
for sheet in wb_ref.sheetnames:
    ws = wb_ref[sheet]
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, 2).value or "").strip().lower()
        contact = str(ws.cell(r, 3).value or "").strip().lower()
        if nom:
            existing[nom] = {"nom": ws.cell(r, 2).value, "contact": contact}

def normalize(s):
    s = s.lower().strip()
    s = re.sub(r'\b(sa|sarl|gmbh|ag|sàrl|cie|groupe)\b', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def is_dup(nom, email=""):
    nb = normalize(nom)
    for ex in existing:
        eb = normalize(ex)
        if len(nb) > 3 and len(eb) > 3:
            if nb == eb or nb in eb or eb in nb:
                return True
        if email:
            ex_contact = str(existing[ex]["contact"] or "").lower().strip()
            if email.lower().strip() == ex_contact:
                return True
    return False

# === GENEVE (APPT annuary + web) ===
geneve = [
    ("Apleona GVA AG", "Genève", "+41228270202", "geneve.realestate.ch@bilfinger.com", "http://www.realestate-ch.apleona.com"),
    ("Argecil", "Genève", "+41227070800", "info@argecil.ch", "http://www.argecil.ch"),
    ("Gérard BAEZNER & Cie SA", "Genève", "+41227070350", "location@regiebaezner.ch", "http://www.regiebaezner.ch"),
    ("Barnes Genève", "Genève", "+41228090690", "", "https://www.barnes-suisse.ch"),
    ("Groupe Bernard Nicod SA", "Genève", "+41227180888", "info@bernard-nicod.ch", "http://www.bernard-nicod.ch"),
    ("Régie Bersier & Cie SA", "Genève", "+41228279010", "regie@bersiersa.ch", "http://www.bersiersa.ch"),
    ("Charles Besuchet", "Genève", "+41227070760", "contact@besuchet.ch", "http://www.besuchet.ch"),
    ("Bordier & Schmidhauser", "Genève", "+41228191119", "location@bordier-schmidhauser.ch", "https://www.bordier-schmidhauser.ch"),
    ("Bory immobilier", "Genève", "+41227081212", "location@bory.ch", "http://www.bory.ch"),
    ("Brolliet SA", "Genève", "+41582013500", "", "http://www.brolliet.ch"),
    ("Edouard Brun & Cie S.A.", "Genève", "+41227181919", "location@regiebrun.ch", "http://www.regiebrun.ch"),
    ("Rodolphe Burger SA", "Genève", "+41227165050", "", "http://www.burger-sa.ch"),
    ("Régie du Centre SA", "Genève", "+41223196555", "location@regieducentre.ch", "http://www.regieducentre.ch"),
    ("COGERIM SCOP", "Genève", "+41225943434", "location@cogerim.ch", "https://www.cogerim.ch"),
    ("Comptoir Immobilier", "Genève", "+41223198888", "locations@comptoir-immo.ch", "http://www.comptoir-immo.ch"),
    ("Régie Foncière SA", "Genève", "+41228170817", "info@regiefonciere.ch", "http://www.regiefonciere.ch"),
    ("Gerofinance | Régie du Rhône GE", "Genève", "+41228090600", "geneve@gerofinance.ch", "https://www.gerofinance.ch"),
    ("Grange & Cie SA", "Genève", "+41227071010", "contact@grange.ch", "http://www.grange.ch"),
    ("Régisseurs du Léman", "Genève", "+41228398400", "info@regisseursduleman.ch", "http://www.regisseursduleman.ch"),
    ("m3 REAL ESTATE", "Genève", "+41228090909", "m3@m-3.com", "http://www.m-3.com"),
    ("Groupe Immobilier du Mail", "Genève", "+41227041900", "gerance@regies.ch", "http://www.regies.ch"),
    ("Moser Vernet & Cie SA", "Genève", "+41228390925", "locations.accueil@moservernet.ch", "https://www.moservernet.ch"),
    ("Naef Immobilier Genève", "Genève", "+41228393939", "geneve@naef.ch", "http://www.naef.ch"),
    ("Pilet & Renaud SA", "Genève", "+41223229270", "", "http://pilet-renaud.ch"),
    ("Les Régisseurs Associés", "Genève", "+41227875000", "info@regisseurs.ch", "http://www.regisseurs.ch"),
    ("Régie du Rhône SA", "Genève", "+41582190000", "geneve@regierhone.ch", "http://www.regierhone.ch"),
    ("ROSSET & Cie SA", "Genève", "+41223393939", "location@rosset.ch", "http://www.rosset.ch"),
    ("Société Privée de Gérance (SPG)", "Genève", "+41588103000", "location@spg.ch", "https://www.spg-rytz.ch"),
    ("Tournier Régie", "Genève", "+41223183070", "regie@tournier.ch", "https://tournier.ch"),
    ("Regie Zimmermann", "Genève", "+41582018500", "geneve@zimmo.ch", "https://www.zimmo.ch"),
    ("Rosset Immobilier", "Genève/Lausanne", "+41223393939", "location@rosset.ch", "https://www.rosset.ch"),
]

# === VAUD (Nyon, La Côte, Lausanne, Yverdon) ===
vaud = [
    ("Régie Marmillod SA", "Lausanne 1010", "+41213312200", "info@regiemarmillod.ch", "https://www.regiemarmillod.ch"),
    ("IMMOSPHERE SA", "Cugy 1053", "+41218004000", "info@regie-immosphere.ch", "https://regie-immosphere.ch"),
    ("At-Home Régie Immobilière SA", "Lausanne 1018", "+41216830202", "info@regieathome.ch", "https://regieathome.ch"),
    ("Misa Gérance", "Lausanne", "+41213100860", "info@misa-gerance.ch", "https://www.misa-gerance.ch"),
    ("Galland & Cie SA", "Lausanne 1003", "+41213102525", "info@regiegalland.ch", "http://www.galland.ch"),
    ("Naef Immobilier Lausanne", "Prilly 1008", "+41213187777", "lausanne@naef.ch", "https://www.naef.ch"),
    ("Stalder Immobilier SA", "Cheseaux", "+41218411044", "info@stalder-immobilier.ch", "https://stalder-immobilier.ch"),
    ("Thomas Régie Foncière SA", "Lausanne", "+41217065454", "info@thomas-regiefonciere.ch", "https://www.thomas-regiefonciere.ch"),
    ("Régie Privée SA", "Lausanne/Genève", "", "info@regieprivee.ch", "https://regieprivee.ch"),
    ("Régimo Lausanne", "Renens 1020", "+41213187389", "", "https://regimo-lausanne.ch"),
    ("Duboux Immobilier (Régie Duboux)", "Lausanne/Riviera", "+41219619830", "riviera@regieduboux.ch", "https://regieduboux.ch"),
    ("Cogestim SA", "Lausanne", "", "", "https://cogestim.ch"),
    ("M&B Gérance Immobilière SA", "Lausanne", "", "", "https://www.mbsa.ch"),
    ("Régie du Croset", "Crissier", "", "", "https://www.regieducroset.ch"),
    ("Regicôte SA", "Nyon 1260", "", "", "https://regicote.ch"),
    ("Naef Immobilier Nyon", "Nyon 1260", "+41229942323", "nyon@naef.ch", "https://www.naef.ch"),
    ("Maillard Immobilier", "Nyon 1260", "", "", "https://www.maillard-immo.ch"),
    ("Ad Valorem Immobilier", "Coppet/Nyon/Morges", "", "", "https://ad-valorem.ch"),
    ("L'Agence Immobilière Nyon (Righetti)", "Nyon 1260", "", "", "https://lagence-immobiliere.ch"),
    ("CFP Immo + Conseils", "Nyon/Gland", "", "", "https://www.cfp-immo.ch"),
    ("Verbel Immobilier", "Nyon 1260", "", "", "https://www.verbel.ch"),
    ("BARNES La Côte", "Coppet/Nyon", "", "", "https://www.barnes-suisse.ch"),
    ("Bernard Nicod Morges", "Morges 1110", "+41218047979", "", "http://www.bernard-nicod.ch"),
    ("Bernard Nicod Lausanne", "Lausanne", "", "", "http://www.bernard-nicod.ch"),
    ("Gerofinance Régie du Rhône Vaud", "Lausanne", "", "", "https://www.gerofinance.ch"),
    ("AMMA Immobilier", "Yverdon 1400", "+41244257070", "info@amma.immo", "https://amma.immo"),
    ("BuybyePME Sàrl", "Perroy 1166", "+41796371324", "", ""),
    ("Procité SA", "Lausanne", "", "", ""),
]

# === FRIBOURG ===
fribourg = [
    ("Régie de Fribourg SA", "Fribourg 1700", "+41263505511", "reception@rfsa.ch", "https://www.rfsa.ch"),
    ("REGIS SA", "Fribourg 1701", "+41263501000", "immobilier@regis-sa.ch", "https://www.regis-sa.ch"),
    ("Gérances Foncières SA", "Fribourg 1700", "+41263225441", "office@gerances-foncieres.ch", "https://www.gerances-foncieres.ch"),
    ("Naef Immobilier Fribourg", "Fribourg", "+41263092888", "", "https://www.naef.ch"),
    ("Rosset Immobilier Fribourg", "Fribourg", "", "", "https://www.rosset.ch"),
    ("Gerofinance Régie du Rhône Bulle", "Bulle 1630", "+41263474790", "location-fribourg@grrsa.ch", "https://www.gerofinance.ch"),
    ("Bernard Nicod Fribourg", "Fribourg", "", "", ""),
]

# === NEUCHATEL ===
neuchatel = [
    ("Gerofinance Régie du Rhône Neuchâtel", "Neuchâtel 2000", "+41327232353", "neuchatel@grrsa.ch", "https://www.gerofinance.ch"),
    ("Naef Immobilier Neuchâtel", "Neuchâtel", "+41327372727", "", "https://www.naef.ch"),
    ("Procité SA", "Neuchâtel 2000", "+41327241111", "", ""),
    ("Régie Privée Neuchâtel", "Neuchâtel", "", "", ""),
]

all_prospects = geneve + vaud + fribourg + neuchatel

# Deduplicate
new_prospects = []
seen_emails = set()
dup_email = 0
in_existing = 0

for nom, lieu, tel, email, site in all_prospects:
    if email and email.lower().strip() in seen_emails:
        dup_email += 1
        continue
    if is_dup(nom, email):
        in_existing += 1
        continue
    if email:
        seen_emails.add(email.lower().strip())
    new_prospects.append((nom, lieu, tel, email, site))

print(f"Sources: {len(all_prospects)}")
print(f"Deja dans ta liste: {in_existing}")
print(f"Doublons email: {dup_email}")
print(f"Nouveaux prospects: {len(new_prospects)}")

# Write Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Regies et Agences"

headers = ["Nom entreprise", "Lieu / Code postal", "Telephone", "Email", "Site web", "Statut", "Canton"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, (nom, lieu, tel, email, site) in enumerate(new_prospects, 2):
    ws.cell(i, 1, nom)
    ws.cell(i, 2, lieu)
    ws.cell(i, 3, tel)
    ws.cell(i, 4, email)
    ws.cell(i, 5, site)
    ws.cell(i, 6, "Nouveau prospect")
    # Guess canton
    if any(c in lieu for c in ["Genève"]) or any(c in lieu for c in ["GE"]):
        canton = "Geneve"
    elif any(c in lieu for c in ["Lausanne", "Nyon", "Morges", "Renens", "Crissier", "Rolle", "Gland", "Aubonne", "Coppet", "Yverdon", "Cheseaux", "Prilly", "Cugy", "Perroy", "Vich"]):
        canton = "Vaud"
    elif any(c in lieu for c in ["Fribourg", "Bulle"]):
        canton = "Fribourg"
    elif "Neuch" in lieu:
        canton = "Neuchatel"
    else:
        canton = "Suisse romande"
    ws.cell(i, 7, canton)

for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(new_prospects) + 2):
        val = ws.cell(row, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[chr(64 + col)].width = max_len + 2

output = r"C:\Users\openc\.openclaw\workspace838360131817\Nouveaux_Prospects_Regies_Agences.xlsx"
wb.save(output)
print(f"Fichier: {output}")

# Stats
from collections import Counter
cantons = Counter()
for _, l, _, _, _ in new_prospects:
    if any(c in l for c in ["Genève"]):
        cantons["Geneve"] += 1
    elif any(c in l for c in ["Lausanne", "Nyon", "Morges", "Renens", "Crissier", "Rolle", "Gland", "Aubonne", "Coppet", "Yverdon", "Cheseaux", "Prilly", "Cugy", "Perroy", "Vich"]):
        cantons["Vaud"] += 1
    elif any(c in l for c in ["Fribourg", "Bulle"]):
        cantons["Fribourg"] += 1
    elif "Neuch" in l:
        cantons["Neuchatel"] += 1
    else:
        cantons["Suisse romande"] += 1

print(f"\nRepartition:")
for c, n in sorted(cantons.items()):
    print(f"  {c}: {n}")
print(f"\nAvec email: {sum(1 for _,_,_,e,_ in new_prospects if e)}")
print(f"Avec telephone: {sum(1 for _,_,t,_,_ in new_prospects if t)}")
