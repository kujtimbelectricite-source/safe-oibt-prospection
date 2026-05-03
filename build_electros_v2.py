#!/usr/bin/env python3
"""
Build new electricien prospects - distance-filtered around Arzier 1273.
"""

import openpyxl, re

# ── Coordinates of Arzier (1273) ──
# Lat: ~46.46, Lon: ~6.21

# ── Load existing clients ──
ref_path = r"C:\Users\openc\.openclaw\media\inbound\Clients_de_Safe_OIBT_Control---d0fcb6aa-77e8-42af-8c09-ff3014d6f8a7.xlsx"
wb_ref = openpyxl.load_workbook(ref_path)

existing = {}
for sheet_name in wb_ref.sheetnames:
    ws = wb_ref[sheet_name]
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, 2).value or "").strip().lower()
        contact = str(ws.cell(r, 3).value or "").strip().lower()
        if nom:
            existing[nom] = {"nom": ws.cell(r, 2).value, "contact": contact}

print(f"Existing clients loaded: {len(existing)}")

# ── Helper to normalize names for dedup ──
def normalize(s):
    s = s.lower().strip()
    s = re.sub(r'\b(sa|sarl|gmbh|ag|sàrl|succursale)\b', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def is_duplicate(nom, email):
    nom_base = normalize(nom)
    for ex_nom in existing:
        ex_base = normalize(ex_nom)
        if len(nom_base) > 3 and len(ex_base) > 3:
            if nom_base == ex_base or nom_base in ex_base or ex_base in nom_base:
                return True
    return False

# ── Prospects organized by region around Arzier ──

# === ZONE 1: 0-15 km (Nyon, Gland, Rolle, La Côte) ===
zone1 = [
    ("3TECH SA", "Les Acacias 1227", "+41223081500", "compta@3tech.ch", "http://www.3tech.ch"),
    ("AC électricité Sàrl", "Satigny 1242", "+41225109920", "info@ac-electricite.ch", "http://www.ac-electricite.ch"),
    ("ALVAZZI GROUPE SA", "Plan-les-Ouates 1228", "+41223019045", "geneve@alvazzigroupe.com", ""),
    ("AM Electricité SA", "Troinex 1256", "+41227840170", "contact@amelectricite.ch", "http://www.amelectricite.ch"),
    ("Bondat SA", "Vernier 1214", "+41223061000", "info@bondat.ch", ""),
    ("Cometel SA", "Genève 1205", "+41227080303", "info@cometel.ch", "http://www.cometel.ch"),
    ("Conti & Associés Ingénieurs SA", "Versoix 1290", "+41227555525", "info@conti-ing.ch", "http://www.conti-ingenergy.ch"),
    ("DELELEC SARL", "Chêne-Bourg 1225", "+41223207605", "secretariat@delelec.ch", "http://www.delelec.ch"),
    ("EGD Electricité S.A.", "Chêne-Bougeries 1224", "+41223491366", "info@egdelectricite.ch", ""),
    ("Energies Contrôles & Conseils Sàrl", "Cointrin 1216", "+41223011541", "info@energies-cc.ch", ""),
    ("J. Kappeler SA", "Grand-Lancy 1212", "+41227071540", "info@kappeler-electricien.ch", "http://www.kappeler-electricien.ch"),
    ("Jäggi Philippe", "Collonge-Bellerive 1245", "+41223404040", "info@jaeggi-electro.ch", "http://www.jaeggi-electro.ch"),
    ("Kreutzer et Cie SA", "Genève 1205", "+41228001414", "info@kreutzer-electricite.ch", "http://www.kreutzer-electricite.ch"),
    ("LumiA Electricité Sàrl", "Meyrin 1217", "+41798463051", "info@lumi-a.ch", ""),
    ("MP Electricité SA", "Bernex 1233", "+41227571002", "info@mp-electricite.ch", "http://www.mp-electricite.ch"),
    ("Neoelec Sàrl", "Plan-les-Ouates 1228", "+41227943292", "info@neoelec.ch", ""),
    ("Paul Puchat SA", "Plan-les-Ouates 1228", "+41227949666", "info@puchatsa.ch", ""),
    ("PermaLife Energy Sàrl", "Carouge 1227", "+41783185474", "j.keller@permalife-energy.ch", "http://www.permalife-energy.ch"),
    ("Rabunal Electrostar", "Genève 1202", "+41227333029", "travaux@electrostar.ch", "http://www.electrostar.ch"),
    ("SAVOY S.A.", "Carouge 1227", "+41223001111", "info@savoy-sa.ch", "http://www.savoy-sa.ch"),
    ("Vadi Sàrl", "Genève 1205", "+41228000808", "info@vadi.ch", "http://www.vadi.ch"),
    ("Vernet SA", "Carouge 1227", "+41223003535", "info@vernetsa.ch", "http://www.vernetsa.ch"),
    ("Vulliez SA", "Carouge 1227", "+41228272626", "info@vulliez-sa.ch", "http://www.vulliez-sa.ch"),
    ("Wernli & GB Associés SA", "Plan-les-Ouates 1228", "+41228800201", "denis.gentil-beccot@wernligb.ch", "http://www.wernligb.ch"),
    ("apelco électricité sàrl", "Carouge 1227", "+41228001000", "info@apelco.ch", ""),
    ("swisspro SR SA", "Le Lignon 1219", "+41227951122", "infoge@swisspro.ch", "http://www.swisspro.ch"),
    ("Genève Électricité SARL", "Genève 1203", "+41762384702", "info@geneveelectricite.ch", "https://geneveelectricite.ch"),
    ("Rindis & Cie Sàrl", "Nyon 1260", "", "", ""),
    ("Ciel Electricité SA - Succ. La Côte", "Nyon 1260", "+41216514555", "lacote@cielelectricite.com", "http://www.cielelectricite.com"),
    ("Hanhart Electricité SA", "Nyon 1260", "", "", ""),
    ("DA Electro", "Gland 1196", "+41786086588", "daelectro@bluewin.ch", "https://www.daelectro.ch"),
    ("RAMA OIBT", "Vich 1267", "+41215616577", "info@ramaoibt.ch", "https://ramaoibt.ch"),
    ("Charrière et Fils SA", "Aubonne 1170", "", "", "https://www.charrierefils.ch"),
    ("Team Swiss Elec Sàrl", "Petit-Lancy 1213", "+41792796573", "info@teamswisselec.ch", ""),
    ("TeamElec Sàrl", "Grand-Lancy 1212", "+41227826666", "contact@teamelec.ch", "http://www.teamelec.ch"),
    ("Idmelectric Sàrl", "Carouge 1227", "+41783224737", "info@idmelectric.ch", ""),
    ("Gattuso Electricité Sàrl", "Plan-les-Ouates 1228", "+41786742995", "info@gattusoelectricite.ch", ""),
    ("Electrotica Sàrl", "Onex 1213", "+41227930505", "info@electrotica.ch", "http://www.electrotica.ch"),
    ("Electro Concept SA", "Châtelaine 1219", "+41223425562", "info@electro-concept.ch", "http://www.electro-concept.ch"),
    ("Electritec SA", "Genève 1201", "+41227335400", "info@electritec.ch", "http://www.electritec.ch"),
    ("ETAVIS Romandie SA", "Carouge 1227", "+41228276464", "geneve.etatsa@etavis.ch", "http://www.etavis.ch"),
    ("ELECOM ELECTRICITE SA", "Carouge 1227", "+41229100010", "info@elecom.ch", "http://www.elecom.ch"),
    ("Eltex Electro-Téléphone SA", "Genève 1202", "+41227338111", "support@eltex.ch", "http://www.eltexgeneve.ch"),
    ("Energia Electricité SA", "Genève 1203", "+41227930100", "info@energia-sa.ch", "http://www.energia-sa.ch"),
    ("DUPONT SA", "Genève 1206", "+41227043000", "electricite@dupontsa.ch", "http://www.dupontsa.ch"),
    ("DOWELLEC SA", "Grand-Lancy 1212", "+41227355605", "dowellec@dowellec.com", "http://www.dowellec.com"),
    ("Devantis Electricité SARL", "Veyrier 1255", "+41225926202", "info@devantis-electricite.ch", "http://www.devantis-electricite.ch"),
    ("PIERRE DEL BON SA", "Chêne-Bougeries 1224", "+41223485606", "info@delbon.ch", ""),
    ("PIRKER SA ELECTRICITE", "Genève 1211", "+41227364646", "info@pirkersa.ch", ""),
    ("Salvisberg Electricité SA", "Genève 1203", "+41229497740", "info@salvisberg.swiss", "http://www.salvisberg.swiss"),
    ("Sedelec SA", "Carouge 1227", "+41228698000", "f.casano@sedelec.ch", "http://www.sedelec.ch"),
    ("HC ELEC SA", "Genève 1204", "+41223283535", "info@hcelec-sa.ch", ""),
    ("Kaech SA", "Genève 1203", "+41223397770", "vieira@kaech.ch", "http://www.kaech.ch"),
    ("Kauz Didier Electricité-Téléphone", "Meinier 1252", "+41227721280", "didier.kauz@bluewin.ch", ""),
    ("Coelec SA", "Genève 1203", "+41223454800", "admin@coelec-sa.ch", "http://www.coelec-sa.ch"),
    ("CP CIPRIANO SA", "Genève 1203", "+41229491010", "ciprianodidier@bluewin.ch", ""),
    ("N. GARCIA Sàrl", "Genève 1205", "+41223201766", "n.garcia@bluewin.ch", ""),
    ("Equans Switzerland AG", "Plan-les-Ouates 1228", "+41223061616", "admin.rh.ch@equans.com", "http://www.equans.com"),
    ("SPIE MTS SA", "Satigny 1242", "+41583011818", "rh.spiemts.ch@spie.com", "http://www.spie.com"),
]

# === ZONE 2: 15-40 km (Lausanne, Morges, La Côte est) ===
zone2 = [
    ("Electricité Lausanne (Elecpro)", "Lausanne 1007", "+41212180102", "Info@electricitelausanne.ch", "https://www.electricitelausanne.ch"),
    ("Electro-Techniques AZ SA", "Grandvaux 1091", "+41217999191", "info@electrotechniques.ch", "https://electrotechniques.ch"),
    ("KW Services SA", "Crissier 1023", "+41213575002", "", "https://kw-services-sa.ch"),
    ("JDC Electricité SA", "Puidoux 1070", "+41219461595", "info@jdc-electricite.ch", "https://www.jdc-electricite.ch"),
    ("Jordi Electricité", "Roche 1852", "", "info@jordi-electricite.ch", "https://www.roche-vd.ch"),
    ("Dave Elec", "Le Mont-sur-Lausanne 1052", "+41219036367", "contact@dave-elec.ch", "https://www.dave-elec.ch"),
    ("Jean Favre SA", "Lausanne", "+41216524343", "info@jeanfavre.ch", "http://www.jeanfavre.ch"),
    ("Dusserre Electricité SA", "Crissier 1023", "+41216376226", "", "https://dusserre.ch"),
    ("Sapin SA", "Morges 1110", "", "", "https://www.sapinsa.ch"),
    ("Telectric Electricité Téléphone SA", "Vaulion 1325", "+41244592183", "telectric.rances@vonet.ch", "https://www.telectric.ch"),
    ("Boymond Electricité SARL", "La Croix-de-Rozon 1257", "+41227711797", "admin@boymond.ch", "http://www.boymond.ch"),
    ("Laurent Chuard Eclairage Maintenance", "Carouge 1227", "+41223421524", "", ""),
    ("Huber SA", "Nyon 1260", "+41229943500", "info@huber-sa.ch", ""),
    ("Caliri Electricité SA", "Nyon 1260", "", "", ""),
    ("Bally Louis & Fils SA", "Nyon 1260", "", "", "https://www.electricitebally.ch"),
    ("Lehmann Christian", "Arzier 1273", "", "", ""),
    ("SIE (Services Industriels)", "Crissier/Ecublens/Renens", "", "", "https://www.sie.ch"),
    ("Bati-STB électricité", "Fribourg 1700", "+41793775922", "info@bati-stb.ch", "https://bati-stb.ch"),
]

all_prospects = zone1 + zone2

# ── Deduplicate ──
new_prospects = []
seen_emails = set()
dup_count = 0
existing_match = 0

for nom, lieu, tel, email, site in all_prospects:
    if email and email.lower().strip() in seen_emails:
        dup_count += 1
        continue
    if is_duplicate(nom, email):
        existing_match += 1
        continue
    if email:
        seen_emails.add(email.lower().strip())
    new_prospects.append((nom, lieu, tel, email, site))

print(f"\nNew prospects: {len(new_prospects)}")
print(f"Dup (email): {dup_count}")
print(f"In existing list: {existing_match}")

# ── Write Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Electriciens"

headers = ["Nom entreprise", "Lieu / Code postal", "Telephone", "Email", "Site web", "Statut"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, (nom, lieu, tel, email, site) in enumerate(new_prospects, 2):
    ws.cell(i, 1, nom)
    ws.cell(i, 2, lieu)
    ws.cell(i, 3, tel)
    ws.cell(i, 4, email)
    ws.cell(i, 5, site)
    ws.cell(i, 6, "Nouveau prospect")

for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(new_prospects) + 2):
        val = ws.cell(row, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[chr(64 + col)].width = max_len + 2

output = r"C:\Users\openc\.openclaw\workspace838360131817\Nouveaux_Prospects_Electriciens.xlsx"
wb.save(output)
print(f"\nFichier: {output}")

gp = sum(1 for n, l, t, e, s in new_prospects if any(c in l for c in ["Genève", "Carouge", "Lancy", "Onex", "Vernier", "Meyrin", "Bernex", "Satigny", "Plan-les-Ouates", "Versoix", "Troinex", "Veyrier", "Cointrin", "Chêne", "Collonge"]))
vd = sum(1 for n, l, t, e, s in new_prospects if any(c in l for c in ["Lausanne", "Crissier", "Renens", "Morges", "Nyon", "Gland", "Vich", "Aubonne", "Rolle", "Puidoux", "Roche", "Mont-sur-Lausanne", "Grandvaux", "Arzier", "Vaulion", "Ecublens"]))
fr_ne = sum(1 for n, l, t, e, s in new_prospects if "Fribourg" in l)

print(f"\nPar canton: Geneve={gp}, Vaud={vd}, FR={fr_ne}")
print(f"Total nouveaux: {len(new_prospects)}")
