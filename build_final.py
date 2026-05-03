#!/usr/bin/env python3
"""
Final comprehensive list of electricien prospects within 100km of Arzier.
Vaud + Geneva + Fribourg + Neuchâtel + Valais (west) + Jura (south).
Filters out existing clients from the reference file.
"""

import openpyxl, re

ref_path = r"C:\Users\openc\.openclaw\media\inbound\Clients_de_Safe_OIBT_Control---d0fcb6aa-77e8-42af-8c09-ff3014d6f8a7.xlsx"
wb_ref = openpyxl.load_workbook(ref_path)

existing = {}
for sheet in wb_ref.sheetnames:
    ws = wb_ref[sheet]
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, 2).value or "").strip().lower()
        contact = str(ws.cell(r, 3).value or "").strip().lower()
        if nom:
            existing[nom] = {"nom": ws.cell(r, 2).value, "contact": contact}

def normalize(s):
    s = s.lower().strip()
    s = re.sub(r'\b(sa|sarl|gmbh|ag|sàrl|succursale)\b', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def is_dup(nom, email):
    nb = normalize(nom)
    for ex in existing:
        eb = normalize(ex)
        if len(nb) > 3 and len(eb) > 3:
            if nb == eb or nb in eb or eb in nb:
                return True
        if email:
            ex_contact = str(existing[ex]["contact"] or "").lower().strip()
            if email.lower().strip() == ex_contact:
                return True
    return False

# === VAUD (all regions) ===
vaud = [
    # Nyon & La Côte (0-20km)
    ("AA Elec Controles SA (aa.elec)", "Rolle", "+41215521075", "aa.elec@bluewin.ch", "https://www.aaeleccontroles.ch"),
    ("Constantin Electricité", "Morges 1110", "+41218241436", "patrick@constantinelectricite.ch", "https://constantinelectricite.ch"),
    ("Bersier Electricité SA", "Préverenges 1028", "+41218031111", "info@bersierelect.ch", "https://www.bersierelect.ch"),
    ("Bally Electricité Yverdon SA", "Yverdon 1400", "", "", "https://edirex.ch"),
    ("Crestoelec Sàrl", "Yverdon 1400", "", "info@crestoelec.ch", "https://www.crestoelec.ch"),
    ("Bärtschi Electricité Sàrl", "Yverdon 1400", "", "", "https://search.ch"),
    ("Jordi SA", "Le Mont-sur-Lausanne 1052", "+41213103925", "courriel@jordisa.ch", "http://www.jordisa.ch"),
    ("Securelec Vaud", "Morges/Nyon", "+41223081620", "securelec@securelec.ch", "https://vaud.securelec.ch"),
    ("VOé (Vaud-Ouest Énergies)", "Yverdon", "", "", "https://www.voe.ch"),
    ("Dépannage Electricien (service)", "Yverdon", "+41783193282", "", "https://depannage-electricien.ch"),
    ("Bally Électricité (electrobally.ch)", "Yverdon", "", "", "https://www.electrobally.ch"),
    ("SIE SA", "Crissier/Ecublens/Renens", "", "", "https://www.sie.ch"),
    ("Dusserre Electricité SA", "Crissier 1023", "+41216376226", "info@dusserre.ch", "https://dusserre.ch"),
    ("Sapin SA", "Morges 1110", "", "", "https://www.sapinsa.ch"),
    ("Telectric (Dany Moffrand)", "Vaulion 1325", "+41244592183", "telectric.rances@vonet.ch", "https://www.telectric.ch"),
    ("Electricité Lausanne - Elecpro", "Lausanne 1007", "+41212180102", "Info@electricitelausanne.ch", "https://www.electricitelausanne.ch"),
    ("KW Services SA", "Crissier 1023", "+41213575002", "", "https://kw-services-sa.ch"),
    ("JDC Electricité SA", "Puidoux 1070", "+41219461595", "info@jdc-electricite.ch", "https://www.jdc-electricite.ch"),
    ("Jordi Electricité", "Roche 1852", "", "info@jordi-electricite.ch", "https://www.roche-vd.ch"),
    ("Dave Elec", "Le Mont-sur-Lausanne 1052", "+41219036367", "contact@dave-elec.ch", "https://www.dave-elec.ch"),
    ("Jean Favre SA", "Lausanne", "+41216524343", "info@jeanfavre.ch", "http://www.jeanfavre.ch"),
    ("Electro-Techniques AZ SA", "Grandvaux/Lausanne", "+41217999191", "info@electrotechniques.ch", "https://electrotechniques.ch"),
    ("RAMA OIBT", "Vich 1267", "+41215616577", "info@ramaoibt.ch", "https://ramaoibt.ch"),
    ("Charrière et Fils SA", "Aubonne 1170", "", "", "https://www.charrierefils.ch"),
    ("DA Electro", "Gland 1196", "+41786086588", "daelectro@bluewin.ch", "https://www.daelectro.ch"),
    ("Ciel Electricité SA - Succ. La Côte", "Nyon 1260", "+41216514555", "lacote@cielelectricite.com", "http://www.cielelectricite.com"),
]

# === GENÈVE ===
geneve = [
    ("3TECH SA", "Les Acacias 1227", "+41223081500", "compta@3tech.ch", "http://www.3tech.ch"),
    ("AC électricité Sàrl", "Satigny 1242", "+41225109920", "info@ac-electricite.ch", "http://www.ac-electricite.ch"),
    ("ALVAZZI GROUPE SA", "Plan-les-Ouates 1228", "+41223019045", "geneve@alvazzigroupe.com", ""),
    ("AM Electricité SA", "Troinex 1256", "+41227840170", "contact@amelectricite.ch", "http://www.amelectricite.ch"),
    ("Bondat SA", "Vernier 1214", "+41223061000", "info@bondat.ch", ""),
    ("Boymond Electricité SARL", "La Croix-de-Rozon 1257", "+41227711797", "admin@boymond.ch", "http://www.boymond.ch"),
    ("Coelec SA", "Genève 1203", "+41223454800", "admin@coelec-sa.ch", "http://www.coelec-sa.ch"),
    ("Cometel SA", "Genève 1205", "+41227080303", "info@cometel.ch", "http://www.cometel.ch"),
    ("Conti & Associés Ingénieurs SA", "Versoix 1290", "+41227555525", "info@conti-ing.ch", "http://www.conti-ingenergy.ch"),
    ("CP CIPRIANO SA", "Genève 1203", "+41229491010", "ciprianodidier@bluewin.ch", ""),
    ("DELELEC SARL", "Chêne-Bourg 1225", "+41223207605", "secretariat@delelec.ch", "http://www.delelec.ch"),
    ("Devantis Electricité SARL", "Veyrier 1255", "+41225926202", "info@devantis-electricite.ch", "http://www.devantis-electricite.ch"),
    ("DOWELLEC SA", "Grand-Lancy 1212", "+41227355605", "dowellec@dowellec.com", "http://www.dowellec.com"),
    ("DUPONT SA", "Genève 1206", "+41227043000", "electricite@dupontsa.ch", "http://www.dupontsa.ch"),
    ("EGD Electricité S.A.", "Chêne-Bougeries 1224", "+41223491366", "info@egdelectricite.ch", ""),
    ("ELECOM ELECTRICITE SA", "Carouge 1227", "+41229100010", "info@elecom.ch", "http://www.elecom.ch"),
    ("Electritec SA", "Genève 1201", "+41227335400", "info@electritec.ch", "http://www.electritec.ch"),
    ("Electro Concept SA", "Châtelaine 1219", "+41223425562", "info@electro-concept.ch", "http://www.electro-concept.ch"),
    ("Electrotica Sàrl", "Onex 1213", "+41227930505", "info@electrotica.ch", "http://www.electrotica.ch"),
    ("Eltex Electro-Téléphone SA", "Genève 1202", "+41227338111", "support@eltex.ch", "http://www.eltexgeneve.ch"),
    ("Energia Electricité SA", "Genève 1203", "+41227930100", "info@energia-sa.ch", "http://www.energia-sa.ch"),
    ("Energies Contrôles & Conseils Sàrl", "Cointrin 1216", "+41223011541", "info@energies-cc.ch", ""),
    ("Equans Switzerland AG", "Plan-les-Ouates 1228", "+41223061616", "admin.rh.ch@equans.com", "http://www.equans.com"),
    ("ETAVIS Romandie SA", "Carouge 1227", "+41228276464", "geneve.etatsa@etavis.ch", "http://www.etavis.ch"),
    ("N. GARCIA Sàrl", "Genève 1205", "+41223201766", "n.garcia@bluewin.ch", ""),
    ("Gattuso Electricité Sàrl", "Plan-les-Ouates 1228", "+41786742995", "info@gattusoelectricite.ch", ""),
    ("HC ELEC SA", "Genève 1204", "+41223283535", "info@hcelec-sa.ch", ""),
    ("Idmelectric Sàrl", "Carouge 1227", "+41783224737", "info@idmelectric.ch", ""),
    ("J. Kappeler SA", "Grand-Lancy 1212", "+41227071540", "info@kappeler-electricien.ch", "http://www.kappeler-electricien.ch"),
    ("Jäggi Philippe", "Collonge-Bellerive 1245", "+41223404040", "info@jaeggi-electro.ch", "http://www.jaeggi-electro.ch"),
    ("JANIN SA", "Veyrier 1255", "+41227844084", "info@janin-sa.ch", "http://www.janin-sa.ch"),
    ("Kaech SA", "Genève 1203", "+41223397770", "vieira@kaech.ch", "http://www.kaech.ch"),
    ("Kauz Didier Electricité-Téléphone", "Meinier 1252", "+41227721280", "didier.kauz@bluewin.ch", ""),
    ("Kreutzer et Cie SA", "Genève 1205", "+41228001414", "info@kreutzer-electricite.ch", "http://www.kreutzer-electricite.ch"),
    ("Laurent Chuard Eclairage Maintenance", "Carouge 1227", "+41223421524", "", ""),
    ("LumiA Electricité Sàrl", "Meyrin 1217", "+41798463051", "info@lumi-a.ch", ""),
    ("MP Electricité SA", "Bernex 1233", "+41227571002", "info@mp-electricite.ch", "http://www.mp-electricite.ch"),
    ("Neoelec Sàrl", "Plan-les-Ouates 1228", "+41227943292", "info@neoelec.ch", ""),
    ("PIERRE DEL BON SA", "Chêne-Bougeries 1224", "+41223485606", "info@delbon.ch", ""),
    ("PIRKER SA ELECTRICITE", "Genève 1211", "+41227364646", "info@pirkersa.ch", ""),
    ("Paul Puchat SA", "Plan-les-Ouates 1228", "+41227949666", "info@puchatsa.ch", ""),
    ("PermaLife Energy Sàrl", "Carouge 1227", "+41783185474", "j.keller@permalife-energy.ch", "http://www.permalife-energy.ch"),
    ("Rabunal Electrostar", "Genève 1202", "+41227333029", "travaux@electrostar.ch", "http://www.electrostar.ch"),
    ("Salvisberg Electricité SA", "Genève 1203", "+41229497740", "info@salvisberg.swiss", "http://www.salvisberg.swiss"),
    ("SAVOY S.A.", "Carouge 1227", "+41223001111", "info@savoy-sa.ch", "http://www.savoy-sa.ch"),
    ("Sedelec SA", "Carouge 1227", "+41228698000", "f.casano@sedelec.ch", "http://www.sedelec.ch"),
    ("SPIE MTS SA", "Satigny 1242", "+41583011818", "rh.spiemts.ch@spie.com", "http://www.spie.com"),
    ("Team Swiss Elec Sàrl", "Petit-Lancy 1213", "+41792796573", "info@teamswisselec.ch", ""),
    ("TeamElec Sàrl", "Grand-Lancy 1212", "+41227826666", "contact@teamelec.ch", "http://www.teamelec.ch"),
    ("Vadi Sàrl", "Genève 1205", "+41228000808", "info@vadi.ch", "http://www.vadi.ch"),
    ("Vernet SA", "Carouge 1227", "+41223003535", "info@vernetsa.ch", "http://www.vernetsa.ch"),
    ("Vulliez SA", "Carouge 1227", "+41228272626", "info@vulliez-sa.ch", "http://www.vulliez-sa.ch"),
    ("Wernli & GB Associés SA", "Plan-les-Ouates 1228", "+41228800201", "denis.gentil-beccot@wernligb.ch", "http://www.wernligb.ch"),
    ("apelco électricité sàrl", "Carouge 1227", "+41228001000", "info@apelco.ch", ""),
    ("swisspro SR SA", "Le Lignon 1219", "+41227951122", "infoge@swisspro.ch", "http://www.swisspro.ch"),
    ("Genève Électricité SARL", "Genève 1203", "+41762384702", "info@geneveelectricite.ch", "https://geneveelectricite.ch"),
]

# === FRIBOURG (within ~100km) ===
fribourg = [
    ("Bati-STB électricité", "Fribourg 1700", "+41793775922", "info@bati-stb.ch", "https://bati-stb.ch"),
    ("Groupe E", "Granges-Paccot/Fribourg", "", "info@groupe-e.ch", "https://www.groupe-e.ch"),
    ("Holderélectric SA", "Fribourg", "", "", ""),
    ("EMF (Entreprise Electrique de Fribourg)", "Fribourg", "", "", ""),
    ("Serelec SA", "Bulle 1630", "", "", ""),
]

# === NEUCHÂTEL (within ~80km) ===
neuchatel = [
    ("Viteos SA", "Neuchâtel", "", "info@viteos.ch", "https://viteos.ch"),
    ("Groupe E/Neuchâtel", "Neuchâtel", "", "", ""),
    ("Energie SA (partenaires)", "La Chaux-de-Fonds", "", "", ""),
]

# === VALAIS OUEST (within ~100km, Martigny/Sion area) ===
valais_ouest = [
    ("ElecConcept Valais SA", "Sion 1950", "", "matthieu.fournier@elecconcept.ch", ""),
    ("Electricité Sédunoise SA", "Sion 1950", "", "electricite@elsed.ch", ""),
    ("Electroalpes Sàrl", "Sion 1950", "", "info@electroalpes.ch", "http://www.electroalpes.ch"),
    ("Electro-Techniques AZ SA (Valais)", "Martigny", "+41217999191", "info@electrotechniques.ch", "https://electrotechniques.ch"),
    ("Electricité M.U.D. Fully Sàrl", "Fully 1926", "", "yannick.dubosson@mud-electricite.ch", ""),
    ("Electricité Pierre-Michel Duay SA", "Orsières 1937", "", "pm.duay@bluewin.ch", "https://www.dpm-electricite.ch"),
    ("Electrochoc Sàrl", "Fey (Nendaz) 1996", "", "electrochoc@netplus.ch", ""),
    ("Energitech Sàrl", "Riddes 1908", "", "c.vouillamoz@energitech.ch", ""),
    ("ESC Electro Service SA", "St-Pierre-de-Clages 1955", "", "info@esc-sa.ch", ""),
    ("ELECTRO VALAIS SA", "Sion 1950", "", "info@electrovs.ch", "http://electro-valais.ch"),
]

# === JURA (partie sud, ~80-100km) ===
jura = [
    ("Viteos SA (Jura nord)", "Delémont", "", "", ""),
]

all_prospects = vaud + geneve + fribourg + neuchatel + valais_ouest + jura

# Deduplicate
new_prospects = []
seen_emails = set()
dup_email = 0
in_existing = 0

for nom, lieu, tel, email, site in all_prospects:
    if email and email.lower().strip() in seen_emails:
        dup_email += 1
        continue
    if is_dup(nom, email):
        in_existing += 1
        continue
    if email:
        seen_emails.add(email.lower().strip())
    new_prospects.append((nom, lieu, tel, email, site))

print(f"Total sources: {len(all_prospects)}")
print(f"Deja dans ta liste: {in_existing}")
print(f"Doublons email: {dup_email}")
print(f"Nouveaux prospects: {len(new_prospects)}")

# Write Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Electriciens"

headers = ["Nom entreprise", "Lieu / Code postal", "Telephone", "Email", "Site web", "Statut", "Region"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, (nom, lieu, tel, email, site) in enumerate(new_prospects, 2):
    ws.cell(i, 1, nom)
    ws.cell(i, 2, lieu)
    ws.cell(i, 3, tel)
    ws.cell(i, 4, email)
    ws.cell(i, 5, site)
    ws.cell(i, 6, "Nouveau prospect")
    # Guess region
    if any(c in lieu for c in ["Genève", "Carouge", "Lancy", "Onex", "Vernier", "Meyrin", "Bernex", "Satigny", "Troinex", "Veyrier", "Cointrin", "Chêne", "Collonge", "Châtelaine", "Le Lignon", "La Croix"]):
        region = "Geneve"
    elif any(c in lieu for c in ["Lausanne", "Crissier", "Morges", "Nyon", "Gland", "Rolle", "Aubonne", "Yverdon", "Vich", "Puidoux", "Roche", "Mont-sur-Lausanne", "Grandvaux", "Arzier", "Vaulion", "Preverenges", "Ecublens", "Renens"]):
        region = "Vaud"
    elif "Fribourg" in lieu or "Bulle" in lieu:
        region = "Fribourg"
    elif "Neuch" in lieu or "La Chaux" in lieu:
        region = "Neuchatel"
    elif any(c in lieu for c in ["Sion", "Valais", "Fully", "Orsieres", "Riddes", "Martigny", "Nendaz"]):
        region = "Valais"
    else:
        region = "Suisse romande"
    ws.cell(i, 7, region)

for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(new_prospects) + 2):
        val = ws.cell(row, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 50))
    ws.column_dimensions[chr(64 + col)].width = max_len + 2

output = r"C:\Users\openc\.openclaw\workspace838360131817\Nouveaux_Prospects_Electriciens.xlsx"
wb.save(output)
print(f"\nFichier: {output}")

# Stats
from collections import Counter
regions = Counter()
for _, l, _, _, _ in new_prospects:
    if any(c in l for c in ["Genève", "Carouge", "Lancy", "Onex", "Vernier", "Meyrin", "Bernex", "Satigny", "Troinex", "Veyrier", "Cointrin", "Chêne", "Collonge", "Châtelaine", "Le Lignon", "La Croix"]):
        regions["Geneve"] += 1
    elif any(c in l for c in ["Lausanne", "Crissier", "Morges", "Nyon", "Gland", "Rolle", "Aubonne", "Yverdon", "Vich", "Puidoux", "Roche", "Mont-sur-Lausanne", "Grandvaux", "Arzier", "Vaulion", "Preverenges", "Ecublens", "Renens"]):
        regions["Vaud"] += 1
    elif "Fribourg" in l:
        regions["Fribourg"] += 1
    elif "Neuch" in l:
        regions["Neuchatel"] += 1
    elif any(c in l for c in ["Sion", "Valais", "Fully", "Orsieres", "Riddes", "Martigny"]):
        regions["Valais"] += 1
    else:
        regions["Autre"] += 1

print(f"\nRepartition:")
for r, c in sorted(regions.items()):
    print(f"  {r}: {c}")
print(f"\nTotal: {len(new_prospects)}")
